import streamlit as st
import requests
from datetime import datetime, timedelta
import pandas as pd
import re
from collections import defaultdict
from io import BytesIO
import base64

# ------------------------------------------------------------
# APP CONFIG
# ------------------------------------------------------------
st.set_page_config(page_title="🎯 Faceless Viral Hunter PRO", layout="wide")
st.title("🎯 Faceless Viral Hunter PRO")
st.markdown("**Reddit Stories, AITA, Horror, Cash Cow, Motivation - FACELESS channels ka king!**")

API_KEY = st.secrets["YOUTUBE_API_KEY"]

SEARCH_URL = "https://www.googleapis.com/youtube/v3/search"
VIDEOS_URL = "https://www.googleapis.com/youtube/v3/videos"
CHANNELS_URL = "https://www.googleapis.com/youtube/v3/channels"
ACTIVITIES_URL = "https://www.googleapis.com/youtube/v3/activities"

# ------------------------------------------------------------
# FACELESS DETECTION KEYWORDS (Enhanced)
# ------------------------------------------------------------
FACELESS_INDICATORS = [
    "stories", "reddit", "aita", "am i the", "horror", "scary", "creepy",
    "nightmare", "revenge", "update", "confession", "askreddit", "tifu",
    "relationship", "cheating", "karma", "tales", "narration", "narrator",
    "motivation", "motivational", "stoic", "stoicism", "wisdom", "quotes",
    "facts", "explained", "documentary", "history", "mystery", "unsolved",
    "crime", "true crime", "case", "cash cow", "compilation", "top 10",
    "top 5", "ranking", "countdown", "best of", "worst of", "gaming",
    "gameplay", "walkthrough", "tutorial", "how to", "guide", "tips",
    "ai voice", "text to speech", "tts", "automated", "no face",
    "anonymous", "faceless", "voice over", "voiceover", "narrated"
]

FACELESS_DESCRIPTION_KEYWORDS = [
    "ai generated", "text to speech", "tts", "voice over", "narration",
    "reddit stories", "scary stories", "horror stories", "true stories",
    "motivation", "stoicism", "self improvement", "cash cow", "automated",
    "compilation", "no face", "faceless", "anonymous channel"
]

PREMIUM_COUNTRIES = {
    'US', 'CA', 'GB', 'AU', 'NZ', 'DE', 'FR', 'IT', 'ES', 'NL', 'BE', 'AT', 'CH',
    'SE', 'NO', 'DK', 'FI', 'IE', 'LU', 'JP', 'KR', 'SG', 'HK'
}

MONETIZATION_COUNTRIES = {
    'US', 'CA', 'GB', 'AU', 'NZ', 'DE', 'FR', 'IT', 'ES', 'NL', 'BE', 'AT', 'CH',
    'SE', 'NO', 'DK', 'FI', 'IE', 'LU', 'JP', 'KR', 'SG', 'HK', 'IN', 'BR', 'MX',
    'AR', 'PL', 'CZ', 'RO', 'GR', 'PT', 'HU', 'TW', 'TH', 'MY', 'ID', 'PH', 'VN',
    'ZA', 'NG', 'EG', 'PK', 'BD', 'RU', 'UA', 'TR', 'SA', 'AE', 'IL', 'CL', 'CO', 'PE'
}

# CPM Rates by Country (estimated USD per 1000 views)
CPM_RATES = {
    'US': 4.0, 'CA': 3.5, 'GB': 3.5, 'AU': 4.0, 'NZ': 3.0,
    'DE': 3.5, 'FR': 2.5, 'IT': 2.0, 'ES': 2.0, 'NL': 3.0,
    'BE': 2.5, 'AT': 3.0, 'CH': 4.5, 'SE': 3.0, 'NO': 4.0,
    'DK': 3.0, 'FI': 2.5, 'IE': 3.0, 'LU': 3.5, 'JP': 2.5,
    'KR': 2.0, 'SG': 2.5, 'HK': 2.0, 'IN': 0.5, 'BR': 0.8,
    'MX': 0.7, 'PH': 0.3, 'ID': 0.4, 'PK': 0.3, 'N/A': 1.0
}

# ------------------------------------------------------------
# PDF GENERATION FUNCTION
# ------------------------------------------------------------
def generate_pdf_report(df, stats):
    """Generate PDF report with clickable links using FPDF"""
    try:
        from fpdf import FPDF
        
        class PDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 16)
                self.cell(0, 10, 'Faceless Viral Hunter PRO Report', 0, 1, 'C')
                self.set_font('Arial', 'I', 10)
                self.cell(0, 5, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'C')
                self.ln(5)
            
            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')
        
        pdf = PDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        
        # Summary Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Summary Statistics', 0, 1)
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 6, f"Total Results: {len(df)}", 0, 1)
        pdf.cell(0, 6, f"Total Views: {df['Views'].sum():,}", 0, 1)
        pdf.cell(0, 6, f"Avg Virality: {df['Virality'].mean():.0f}/day", 0, 1)
        pdf.cell(0, 6, f"Monetized Channels: {len(df[df['MonetizationScore'] >= 70])}", 0, 1)
        pdf.ln(10)
        
        # Results Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Top Channels Found', 0, 1)
        
        for idx, row in df.head(50).iterrows():  # Top 50 for PDF
            pdf.set_font('Arial', 'B', 11)
            
            # Clean title for PDF
            title = row['Title'][:60] + "..." if len(row['Title']) > 60 else row['Title']
            title = title.encode('latin-1', 'replace').decode('latin-1')
            
            pdf.cell(0, 8, f"{idx+1}. {title}", 0, 1)
            
            pdf.set_font('Arial', '', 9)
            channel = row['Channel'].encode('latin-1', 'replace').decode('latin-1')
            pdf.cell(0, 5, f"Channel: {channel}", 0, 1)
            pdf.cell(0, 5, f"Views: {row['Views']:,} | Subs: {row['Subs']:,} | Videos: {row['TotalVideos']}", 0, 1)
            pdf.cell(0, 5, f"Virality: {row['Virality']:,}/day | Engagement: {row['Engagement%']}%", 0, 1)
            pdf.cell(0, 5, f"Monetization: {row['MonetizationStatus']} ({row['MonetizationScore']}%)", 0, 1)
            pdf.cell(0, 5, f"Upload Frequency: {row['UploadSchedule']}", 0, 1)
            
            # Add clickable links
            pdf.set_text_color(0, 0, 255)
            pdf.set_font('Arial', 'U', 9)
            pdf.cell(0, 5, f"Video: {row['Link']}", 0, 1, link=row['Link'])
            pdf.cell(0, 5, f"Channel: {row['ChannelLink']}", 0, 1, link=row['ChannelLink'])
            pdf.set_text_color(0, 0, 0)
            pdf.set_font('Arial', '', 9)
            
            pdf.ln(5)
            
            # Add new page if needed
            if pdf.get_y() > 250:
                pdf.add_page()
        
        return pdf.output(dest='S').encode('latin-1')
    
    except ImportError:
        return None


def generate_html_report(df, stats):
    """Generate HTML report with clickable links (alternative to PDF)"""
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Faceless Viral Hunter PRO Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
            .header {{ text-align: center; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 10px; }}
            .summary {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 20px 0; }}
            .stat-box {{ background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
            .stat-box h3 {{ margin: 0; color: #667eea; font-size: 24px; }}
            .stat-box p {{ margin: 5px 0 0 0; color: #666; }}
            .card {{ background: white; margin: 15px 0; padding: 20px; border-radius: 10px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
            .card-header {{ display: flex; justify-content: space-between; align-items: center; }}
            .card-title {{ font-size: 16px; font-weight: bold; color: #333; margin-bottom: 10px; }}
            .card-channel {{ color: #667eea; text-decoration: none; font-weight: bold; }}
            .card-stats {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin: 10px 0; }}
            .stat {{ text-align: center; padding: 10px; background: #f8f9fa; border-radius: 5px; }}
            .stat-value {{ font-size: 18px; font-weight: bold; color: #333; }}
            .stat-label {{ font-size: 11px; color: #666; }}
            .monetized {{ background: #d4edda; color: #155724; padding: 5px 10px; border-radius: 5px; display: inline-block; }}
            .possibly {{ background: #fff3cd; color: #856404; padding: 5px 10px; border-radius: 5px; display: inline-block; }}
            .not-monetized {{ background: #f8d7da; color: #721c24; padding: 5px 10px; border-radius: 5px; display: inline-block; }}
            .links {{ margin-top: 10px; }}
            .links a {{ display: inline-block; margin-right: 15px; padding: 8px 15px; background: #667eea; color: white; text-decoration: none; border-radius: 5px; }}
            .links a:hover {{ background: #5a6fd6; }}
            .thumbnail {{ width: 120px; height: 68px; border-radius: 5px; object-fit: cover; }}
            .faceless-yes {{ color: #28a745; }}
            .faceless-maybe {{ color: #ffc107; }}
            @media print {{
                body {{ margin: 0; }}
                .card {{ page-break-inside: avoid; }}
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🎯 Faceless Viral Hunter PRO Report</h1>
            <p>Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        </div>
        
        <div class="summary">
            <div class="stat-box">
                <h3>{len(df)}</h3>
                <p>Total Channels</p>
            </div>
            <div class="stat-box">
                <h3>{df['Views'].sum():,}</h3>
                <p>Total Views</p>
            </div>
            <div class="stat-box">
                <h3>{df['Virality'].mean():.0f}/day</h3>
                <p>Avg Virality</p>
            </div>
            <div class="stat-box">
                <h3>{len(df[df['MonetizationScore'] >= 70])}</h3>
                <p>Monetized</p>
            </div>
        </div>
        
        <h2>📊 Channel Results</h2>
    """
    
    for idx, row in df.iterrows():
        monetization_class = "monetized" if row['MonetizationScore'] >= 70 else ("possibly" if row['MonetizationScore'] >= 50 else "not-monetized")
        faceless_class = "faceless-yes" if row['Faceless'] == "YES" else "faceless-maybe"
        
        html += f"""
        <div class="card">
            <div class="card-header">
                <div>
                    <div class="card-title">{idx+1}. {row['Title'][:80]}{'...' if len(row['Title']) > 80 else ''}</div>
                    <a href="{row['ChannelLink']}" target="_blank" class="card-channel">📺 {row['Channel']}</a>
                    <span style="margin-left: 10px;">🌍 {row['Country']} | 📅 Created: {row['ChCreated']}</span>
                </div>
                <img src="{row['Thumb']}" class="thumbnail" alt="Thumbnail">
            </div>
            
            <div class="card-stats">
                <div class="stat">
                    <div class="stat-value">{row['Views']:,}</div>
                    <div class="stat-label">👁️ Views</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{row['Subs']:,}</div>
                    <div class="stat-label">👥 Subscribers</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{row['TotalVideos']}</div>
                    <div class="stat-label">🎬 Total Videos</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{row['Virality']:,}/day</div>
                    <div class="stat-label">🔥 Virality</div>
                </div>
            </div>
            
            <div style="margin: 10px 0;">
                <span class="{monetization_class}">{row['MonetizationStatus']} ({row['MonetizationScore']}%)</span>
                <span style="margin-left: 10px;">⏰ {row['UploadSchedule']}</span>
                <span style="margin-left: 10px;" class="{faceless_class}">{'✅ Faceless' if row['Faceless'] == 'YES' else '🤔 Maybe Faceless'} ({row['FacelessScore']}%)</span>
            </div>
            
            <div style="font-size: 12px; color: #666;">
                👍 {row['Likes']:,} likes | 💬 {row['Comments']:,} comments | 
                ⏱️ {row['DurationStr']} ({row['Type']}) | 📤 Uploaded: {row['Uploaded']} |
                🔑 Keyword: {row['Keyword']}
            </div>
            
            <div class="links">
                <a href="{row['Link']}" target="_blank">▶️ Watch Video</a>
                <a href="{row['ChannelLink']}" target="_blank">📺 View Channel</a>
            </div>
        </div>
        """
    
    html += """
    </body>
    </html>
    """
    
    return html


def generate_excel_report(df):
    """Generate Excel report with formatting and hyperlinks"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Viral Channels"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        
        # Select columns for Excel
        excel_cols = ["Title", "Channel", "Views", "Subs", "TotalVideos", "Virality", 
                      "Engagement%", "MonetizationStatus", "MonetizationScore", 
                      "UploadSchedule", "Country", "Faceless", "FacelessScore",
                      "Uploaded", "ChCreated", "Link", "ChannelLink"]
        
        df_excel = df[excel_cols].copy()
        
        # Write headers
        for col_idx, header in enumerate(excel_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, row in enumerate(df_excel.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Make links clickable
                if excel_cols[col_idx-1] == "Link":
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
                elif excel_cols[col_idx-1] == "ChannelLink":
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    except ImportError:
        return None


def estimate_revenue(views, country, video_count):
    """Estimate channel revenue based on views and CPM"""
    cpm = CPM_RATES.get(country, 1.0)
    # Assume 55% of views are monetized (ads not on all videos)
    monetized_views = views * 0.55
    # Calculate revenue
    revenue = (monetized_views / 1000) * cpm
    monthly_revenue = revenue / max((video_count / 30), 1) if video_count > 0 else 0
    return round(revenue, 2), round(monthly_revenue, 2)


def calculate_growth_potential(row):
    """Calculate growth potential score based on multiple factors"""
    score = 0
    
    # Virality factor (high views per day = growing)
    if row['Virality'] > 5000:
        score += 30
    elif row['Virality'] > 2000:
        score += 20
    elif row['Virality'] > 500:
        score += 10
    
    # Engagement factor
    if row['Engagement%'] > 5:
        score += 25
    elif row['Engagement%'] > 2:
        score += 15
    elif row['Engagement%'] > 1:
        score += 10
    
    # Sub to view ratio (viral potential)
    if row['SubViewRatio'] > 5:
        score += 20
    elif row['SubViewRatio'] > 2:
        score += 10
    
    # Upload consistency
    if row['UploadsPerWeek'] >= 3:
        score += 15
    elif row['UploadsPerWeek'] >= 1:
        score += 10
    
    # Young channel bonus
    try:
        created_year = int(row['ChCreated'][:4])
        if created_year >= 2024:
            score += 10
    except:
        pass
    
    return min(score, 100)


def detect_niche(title, channel_name, keyword):
    """Auto-detect the niche category"""
    text = f"{title} {channel_name} {keyword}".lower()
    
    niches = {
        "Reddit Stories": ["reddit", "aita", "am i the", "tifu", "entitled", "revenge", "malicious"],
        "Horror/Scary": ["horror", "scary", "creepy", "nightmare", "paranormal", "ghost", "terror"],
        "True Crime": ["true crime", "crime", "murder", "case", "investigation", "unsolved"],
        "Motivation": ["motivation", "stoic", "stoicism", "mindset", "discipline", "sigma", "self improvement"],
        "Facts/Education": ["facts", "explained", "documentary", "history", "science", "top 10"],
        "Gaming": ["gaming", "gameplay", "walkthrough", "lets play", "gamer"],
        "Compilation": ["compilation", "best of", "funny", "fails", "moments"],
        "Mystery": ["mystery", "mysteries", "unsolved", "conspiracy", "strange"]
    }
    
    for niche, keywords in niches.items():
        if any(kw in text for kw in keywords):
            return niche
    
    return "Other"


# ------------------------------------------------------------
# SIDEBAR - Enhanced Settings
# ------------------------------------------------------------
st.sidebar.header("⚙️ Advanced Settings")

with st.sidebar.expander("📅 Time Filters", expanded=True):
    days = st.slider("Videos from last X days", 1, 90, 14)
    channel_age = st.selectbox(
        "Channel Created After",
        ["2025", "2024", "2023", "2022", "Any"],
        index=1
    )

with st.sidebar.expander("📊 View Filters", expanded=True):
    min_views = st.number_input("Min Views", min_value=1000, value=10000, step=1000)
    max_views = st.number_input("Max Views (0=No Limit)", min_value=0, value=0, step=10000)
    min_virality = st.slider("Min Virality Score (Views/Day)", 0, 10000, 500)

with st.sidebar.expander("👥 Subscriber Filters", expanded=True):
    min_subs = st.number_input("Min Subscribers", min_value=0, value=100)
    max_subs = st.number_input("Max Subscribers", min_value=0, value=500000)

with st.sidebar.expander("🎬 Video Type", expanded=True):
    video_type = st.selectbox("Video Duration", ["All", "Long (5min+)", "Medium (1-5min)", "Shorts (<1min)"])
    
with st.sidebar.expander("🎯 Faceless Detection", expanded=True):
    faceless_only = st.checkbox("Only Faceless Channels", value=True)
    faceless_strictness = st.select_slider(
        "Detection Strictness",
        options=["Relaxed", "Normal", "Strict"],
        value="Normal"
    )

with st.sidebar.expander("💰 Monetization Filter", expanded=False):
    monetized_only = st.checkbox("Only Likely Monetized Channels", value=False)
    min_upload_frequency = st.slider("Min Uploads per Week", 0, 14, 0)

with st.sidebar.expander("🌍 Region Filters", expanded=False):
    premium_only = st.checkbox("Only Premium CPM Countries", value=False)
    search_regions = st.multiselect(
        "Search in Regions",
        ["US", "GB", "CA", "AU", "IN", "PH"],
        default=["US"]
    )

with st.sidebar.expander("🔍 Search Settings", expanded=False):
    search_orders = st.multiselect(
        "Search Order (Multiple = More Results)",
        ["viewCount", "relevance", "date", "rating"],
        default=["viewCount", "relevance"]
    )
    results_per_keyword = st.slider("Results per keyword", 50, 150, 100)
    use_pagination = st.checkbox("Use Pagination (More Results)", value=True)

with st.sidebar.expander("📤 Export Settings", expanded=False):
    export_format = st.multiselect(
        "Export Formats",
        ["CSV", "Excel", "HTML Report", "PDF Report"],
        default=["CSV", "HTML Report"]
    )

# ------------------------------------------------------------
# KEYWORDS INPUT
# ------------------------------------------------------------
st.markdown("### 🔑 Keywords / Titles")

default_keywords = """reddit stories
aita
am i the asshole
reddit relationship advice
reddit cheating stories
true horror stories
scary stories
mr nightmare type
creepypasta
pro revenge reddit
nuclear revenge
malicious compliance
entitled parents
choosing beggars
tifu reddit
best reddit posts
askreddit
reddit updates
relationship drama
motivation
stoicism
stoic quotes
self improvement
marcus aurelius
dark psychology
sigma mindset
cash cow
top 10 facts
explained documentary
true crime
unsolved mysteries
conspiracy theories
history facts
scary mysteries
creepy compilations"""

keyword_input = st.text_area(
    "Enter Keywords (One per line - More keywords = More results)",
    height=300,
    value=default_keywords
)

# Quick keyword templates
col1, col2, col3, col4 = st.columns(4)
with col1:
    if st.button("📖 Reddit Niche"):
        st.session_state.keywords = "reddit stories\naita\nam i the asshole\npro revenge\nnuclear revenge\nmalicious compliance\nentitled parents\nreddit updates\nreddit drama"
with col2:
    if st.button("👻 Horror Niche"):
        st.session_state.keywords = "true horror stories\nscary stories\ncreepypasta\nmr nightmare\nhorror narration\ncreepy stories\nparanormal stories\ntrue scary"
with col3:
    if st.button("💪 Motivation Niche"):
        st.session_state.keywords = "stoicism\nmotivation\nself improvement\nmarcus aurelius\nsigma mindset\ndark psychology\nmindset\ndiscipline"
with col4:
    if st.button("📺 Cash Cow"):
        st.session_state.keywords = "top 10\nfacts about\nexplained\ndocumentary\ntrue crime\nmysteries\nconspiracy\nhistory facts"

# ------------------------------------------------------------
# HELPER FUNCTIONS
# ------------------------------------------------------------
def fetch_json(url, params, retries=2):
    for attempt in range(retries):
        try:
            resp = requests.get(url, params=params, timeout=30)
            if resp.status_code == 200:
                return resp.json()
            if "quotaExceeded" in resp.text:
                return "QUOTA"
            if resp.status_code == 403:
                return "QUOTA"
        except requests.exceptions.Timeout:
            if attempt < retries - 1:
                continue
            return None
        except Exception as e:
            if attempt < retries - 1:
                continue
            return None
    return None


def parse_duration(duration):
    if not duration:
        return 0
    total = 0
    matches = re.findall(r"(\d+)([HMS])", duration)
    for value, unit in matches:
        if unit == "H":
            total += int(value) * 3600
        elif unit == "M":
            total += int(value) * 60
        elif unit == "S":
            total += int(value)
    return total


def calculate_virality_score(views, published_at):
    try:
        pub_date = datetime.strptime(published_at[:19], "%Y-%m-%dT%H:%M:%S")
        days_since = max((datetime.utcnow() - pub_date).days, 1)
        return round(views / days_since, 2)
    except:
        return 0


def calculate_engagement_rate(views, likes, comments):
    if views == 0:
        return 0
    engagement = ((likes + comments * 2) / views) * 100
    return round(engagement, 2)


def calculate_upload_frequency(created_date, total_videos):
    try:
        if not created_date or total_videos == 0:
            return 0, 0, "N/A"
        
        created = datetime.strptime(created_date[:19], "%Y-%m-%dT%H:%M:%S")
        days_active = max((datetime.utcnow() - created).days, 1)
        weeks_active = max(days_active / 7, 1)
        months_active = max(days_active / 30, 1)
        
        uploads_per_week = round(total_videos / weeks_active, 2)
        uploads_per_month = round(total_videos / months_active, 2)
        
        if uploads_per_week >= 7:
            schedule = f"🔥 Daily+ ({uploads_per_week:.1f}/week)"
        elif uploads_per_week >= 3:
            schedule = f"📈 Very Active ({uploads_per_week:.1f}/week)"
        elif uploads_per_week >= 1:
            schedule = f"✅ Regular ({uploads_per_week:.1f}/week)"
        elif uploads_per_week >= 0.5:
            schedule = f"📅 Bi-weekly ({uploads_per_week:.1f}/week)"
        elif uploads_per_week >= 0.25:
            schedule = f"📆 Monthly ({uploads_per_month:.1f}/month)"
        else:
            schedule = f"⏸️ Inactive ({uploads_per_month:.1f}/month)"
        
        return uploads_per_week, uploads_per_month, schedule
        
    except Exception as e:
        return 0, 0, "N/A"


def check_monetization_status(channel_data):
    reasons = []
    score = 0
    
    subs = channel_data.get("subs", 0)
    total_videos = channel_data.get("video_count", 0)
    created = channel_data.get("created", "")
    country = channel_data.get("country", "N/A")
    total_views = channel_data.get("total_views", 0)
    
    if subs >= 1000:
        score += 30
        reasons.append(f"✅ {subs:,} subs (1K+ met)")
    elif subs >= 500:
        score += 10
        reasons.append(f"⏳ {subs:,} subs (close to 1K)")
    else:
        reasons.append(f"❌ {subs:,} subs (needs 1K)")
    
    if created:
        try:
            created_date = datetime.strptime(created[:19], "%Y-%m-%dT%H:%M:%S")
            days_old = (datetime.utcnow() - created_date).days
            
            if days_old >= 365:
                score += 20
                reasons.append(f"✅ {days_old} days old (1yr+)")
            elif days_old >= 30:
                score += 15
                reasons.append(f"✅ {days_old} days old (30d+ met)")
            else:
                reasons.append(f"❌ {days_old} days old (needs 30d)")
        except:
            pass
    
    if country in MONETIZATION_COUNTRIES:
        score += 15
        if country in PREMIUM_COUNTRIES:
            reasons.append(f"✅ {country} (Premium CPM)")
        else:
            reasons.append(f"✅ {country} (Eligible)")
    elif country == "N/A":
        score += 5
        reasons.append("⚠️ Country unknown")
    else:
        reasons.append(f"❌ {country} (May not be eligible)")
    
    estimated_watch_hours = (total_views * 3.2) / 60
    if estimated_watch_hours >= 4000:
        score += 25
        reasons.append(f"✅ ~{estimated_watch_hours:,.0f} est. watch hrs")
    elif estimated_watch_hours >= 2000:
        score += 15
        reasons.append(f"⏳ ~{estimated_watch_hours:,.0f} est. watch hrs")
    else:
        reasons.append(f"❓ ~{estimated_watch_hours:,.0f} est. watch hrs")
    
    if total_videos >= 50:
        score += 10
        reasons.append(f"✅ {total_videos} videos (consistent)")
    elif total_videos >= 20:
        score += 5
        reasons.append(f"📹 {total_videos} videos")
    else:
        reasons.append(f"📹 {total_videos} videos (low)")
    
    if score >= 70:
        status = "🟢 LIKELY MONETIZED"
        confidence = "High"
    elif score >= 50:
        status = "🟡 POSSIBLY MONETIZED"
        confidence = "Medium"
    elif score >= 30:
        status = "🟠 CLOSE TO MONETIZATION"
        confidence = "Low"
    else:
        status = "🔴 NOT MONETIZED"
        confidence = "Very Low"
    
    return status, confidence, score, reasons


def detect_faceless_advanced(channel_data, strictness="Normal"):
    reasons = []
    score = 0
    
    profile_url = channel_data.get("profile", "")
    banner_url = channel_data.get("banner", "")
    channel_name = channel_data.get("name", "").lower()
    description = channel_data.get("description", "").lower()
    
    if "default.jpg" in profile_url or "s88-c-k-c0x00ffffff-no-rj" in profile_url:
        score += 30
        reasons.append("Default profile pic")
    
    if not banner_url:
        score += 20
        reasons.append("No banner")
    
    name_matches = sum(1 for kw in FACELESS_INDICATORS if kw in channel_name)
    if name_matches >= 1:
        score += min(name_matches * 15, 30)
        reasons.append(f"Name matches ({name_matches} keywords)")
    
    desc_matches = sum(1 for kw in FACELESS_DESCRIPTION_KEYWORDS if kw in description)
    if desc_matches >= 1:
        score += min(desc_matches * 10, 25)
        reasons.append(f"Description matches ({desc_matches} keywords)")
    
    ai_patterns = ["ai", "voice", "narrator", "stories", "compilation", "facts", "top"]
    ai_matches = sum(1 for p in ai_patterns if p in channel_name)
    if ai_matches >= 2:
        score += 15
        reasons.append("AI/Compilation pattern")
    
    if channel_data.get("custom_url") is None:
        score += 5
        reasons.append("No custom URL")
    
    thresholds = {"Relaxed": 20, "Normal": 35, "Strict": 55}
    threshold = thresholds.get(strictness, 35)
    
    is_faceless = score >= threshold
    confidence = min(score, 100)
    
    return is_faceless, confidence, reasons


def get_video_type_label(duration):
    if duration < 60:
        return "Shorts"
    elif duration < 300:
        return "Medium"
    else:
        return "Long"


def format_number(num):
    if num >= 1000000:
        return f"{num/1000000:.1f}M"
    elif num >= 1000:
        return f"{num/1000:.1f}K"
    else:
        return str(num)


def batch_fetch_channels(channel_ids, api_key, cache):
    new_ids = [cid for cid in channel_ids if cid not in cache]
    
    if not new_ids:
        return cache
    
    for i in range(0, len(new_ids), 50):
        batch = new_ids[i:i+50]
        params = {
            "part": "snippet,statistics,brandingSettings,status",
            "id": ",".join(batch),
            "key": api_key
        }
        
        data = fetch_json(CHANNELS_URL, params)
        if data == "QUOTA":
            return "QUOTA"
        if not data:
            continue
            
        for c in data.get("items", []):
            sn = c["snippet"]
            stats = c["statistics"]
            brand = c.get("brandingSettings", {})
            status = c.get("status", {})
            brand_img = brand.get("image", {})
            brand_ch = brand.get("channel", {})
            
            profile = sn.get("thumbnails", {}).get("default", {}).get("url", "")
            banner = brand_img.get("bannerExternalUrl", "")
            
            cache[c["id"]] = {
                "name": sn.get("title", ""),
                "subs": int(stats.get("subscriberCount", 0)),
                "total_views": int(stats.get("viewCount", 0)),
                "video_count": int(stats.get("videoCount", 0)),
                "created": sn.get("publishedAt", ""),
                "country": sn.get("country", "N/A"),
                "description": sn.get("description", ""),
                "profile": profile,
                "banner": banner,
                "custom_url": sn.get("customUrl"),
                "keywords": brand_ch.get("keywords", ""),
                "is_linked": status.get("isLinked", False),
                "long_uploads_status": status.get("longUploadsStatus", ""),
                "made_for_kids": status.get("madeForKids", False)
            }
    
    return cache


def search_videos_with_pagination(keyword, params, api_key, max_pages=2):
    all_items = []
    next_token = None
    
    for page in range(max_pages):
        search_params = params.copy()
        search_params["key"] = api_key
        
        if next_token:
            search_params["pageToken"] = next_token
        
        data = fetch_json(SEARCH_URL, search_params)
        
        if data == "QUOTA":
            return "QUOTA"
        if not data:
            break
            
        items = data.get("items", [])
        all_items.extend(items)
        
        next_token = data.get("nextPageToken")
        if not next_token:
            break
    
    return all_items


# ------------------------------------------------------------
# MAIN ACTION
# ------------------------------------------------------------
if st.button("🚀 HUNT FACELESS VIRAL VIDEOS", type="primary", use_container_width=True):
    
    if not keyword_input.strip():
        st.error("⚠️ Keywords daal do bhai!")
        st.stop()
    
    keywords = [kw.strip() for line in keyword_input.splitlines() 
                for kw in line.split(",") if kw.strip()]
    keywords = list(dict.fromkeys(keywords))
    
    all_results = []
    channel_cache = {}
    seen_videos = set()
    
    published_after = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
    
    total_ops = len(keywords) * len(search_orders) * len(search_regions)
    current_op = 0
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    stats = {
        "total_searched": 0,
        "passed_views": 0,
        "passed_subs": 0,
        "passed_age": 0,
        "passed_faceless": 0,
        "final": 0
    }
    
    # MAIN SEARCH LOOP
    for kw in keywords:
        for order in search_orders:
            for region in search_regions:
                current_op += 1
                progress_bar.progress(current_op / total_ops)
                status_text.markdown(f"🔍 **Searching:** `{kw}` | Order: `{order}` | Region: `{region}`")
                
                search_params = {
                    "part": "snippet",
                    "q": kw,
                    "type": "video",
                    "order": order,
                    "publishedAfter": published_after,
                    "maxResults": 50,
                    "regionCode": region,
                    "relevanceLanguage": "en",
                    "safeSearch": "none"
                }
                
                if use_pagination:
                    items = search_videos_with_pagination(kw, search_params, API_KEY, max_pages=2)
                else:
                    data = fetch_json(SEARCH_URL, {**search_params, "key": API_KEY})
                    items = data.get("items", []) if data and data != "QUOTA" else []
                
                if items == "QUOTA":
                    st.error("❌ API Quota khatam! Kal try karo ya API key change karo.")
                    st.stop()
                
                if not items:
                    continue
                
                stats["total_searched"] += len(items)
                
                new_items = []
                for item in items:
                    vid = item.get("id", {}).get("videoId")
                    if vid and vid not in seen_videos:
                        seen_videos.add(vid)
                        new_items.append(item)
                
                if not new_items:
                    continue
                
                video_ids = [i["id"]["videoId"] for i in new_items if "videoId" in i.get("id", {})]
                channel_ids = {i["snippet"]["channelId"] for i in new_items}
                
                video_stats = {}
                for i in range(0, len(video_ids), 50):
                    batch = video_ids[i:i+50]
                    params = {
                        "part": "statistics,contentDetails",
                        "id": ",".join(batch),
                        "key": API_KEY
                    }
                    vid_data = fetch_json(VIDEOS_URL, params)
                    
                    if vid_data == "QUOTA":
                        st.error("❌ API Quota khatam!")
                        st.stop()
                    
                    if vid_data:
                        for v in vid_data.get("items", []):
                            dur_sec = parse_duration(v["contentDetails"].get("duration", ""))
                            s = v.get("statistics", {})
                            video_stats[v["id"]] = {
                                "views": int(s.get("viewCount", 0)),
                                "likes": int(s.get("likeCount", 0)),
                                "comments": int(s.get("commentCount", 0)),
                                "duration": dur_sec
                            }
                
                result = batch_fetch_channels(channel_ids, API_KEY, channel_cache)
                if result == "QUOTA":
                    st.error("❌ API Quota khatam!")
                    st.stop()
                channel_cache = result
                
                for item in new_items:
                    sn = item["snippet"]
                    vid = item["id"].get("videoId")
                    if not vid:
                        continue
                    
                    cid = sn["channelId"]
                    v_stats = video_stats.get(vid, {})
                    ch = channel_cache.get(cid, {})
                    
                    views = v_stats.get("views", 0)
                    likes = v_stats.get("likes", 0)
                    comments = v_stats.get("comments", 0)
                    duration = v_stats.get("duration", 0)
                    subs = ch.get("subs", 0)
                    total_videos = ch.get("video_count", 0)
                    total_channel_views = ch.get("total_views", 0)
                    
                    if views < min_views:
                        continue
                    if max_views > 0 and views > max_views:
                        continue
                    stats["passed_views"] += 1
                    
                    if not (min_subs <= subs <= max_subs):
                        continue
                    stats["passed_subs"] += 1
                    
                    if channel_age != "Any":
                        created_year = int(ch.get("created", "2000")[:4]) if ch.get("created") else 2000
                        if created_year < int(channel_age):
                            continue
                    stats["passed_age"] += 1
                    
                    if faceless_only:
                        is_faceless, confidence, reasons = detect_faceless_advanced(ch, faceless_strictness)
                        if not is_faceless:
                            continue
                    else:
                        is_faceless, confidence, reasons = detect_faceless_advanced(ch, faceless_strictness)
                    stats["passed_faceless"] += 1
                    
                    country = ch.get("country", "N/A")
                    if premium_only and country not in PREMIUM_COUNTRIES:
                        continue
                    
                    vtype = get_video_type_label(duration)
                    if video_type == "Long (5min+)" and duration < 300:
                        continue
                    if video_type == "Medium (1-5min)" and (duration < 60 or duration >= 300):
                        continue
                    if video_type == "Shorts (<1min)" and duration >= 60:
                        continue
                    
                    virality = calculate_virality_score(views, sn["publishedAt"])
                    
                    if virality < min_virality:
                        continue
                    
                    engagement = calculate_engagement_rate(views, likes, comments)
                    sub_view_ratio = round(views / max(subs, 1), 2)
                    
                    uploads_per_week, uploads_per_month, schedule_desc = calculate_upload_frequency(
                        ch.get("created", ""), total_videos
                    )
                    
                    if min_upload_frequency > 0 and uploads_per_week < min_upload_frequency:
                        continue
                    
                    monetization_status, monetization_confidence, monetization_score, monetization_reasons = check_monetization_status(ch)
                    
                    if monetized_only and monetization_score < 50:
                        continue
                    
                    # NEW: Calculate revenue and growth potential
                    est_revenue, monthly_revenue = estimate_revenue(total_channel_views, country, total_videos)
                    niche = detect_niche(sn["title"], sn["channelTitle"], kw)
                    
                    stats["final"] += 1
                    
                    result_data = {
                        "Title": sn["title"],
                        "Channel": sn["channelTitle"],
                        "ChannelID": cid,
                        "Subs": subs,
                        "TotalVideos": total_videos,
                        "TotalChannelViews": total_channel_views,
                        "UploadsPerWeek": uploads_per_week,
                        "UploadsPerMonth": uploads_per_month,
                        "UploadSchedule": schedule_desc,
                        "MonetizationStatus": monetization_status,
                        "MonetizationScore": monetization_score,
                        "MonetizationReasons": " | ".join(monetization_reasons),
                        "EstRevenue": est_revenue,
                        "MonthlyRevenue": monthly_revenue,
                        "Niche": niche,
                        "Views": views,
                        "Likes": likes,
                        "Comments": comments,
                        "Virality": virality,
                        "Engagement%": engagement,
                        "SubViewRatio": sub_view_ratio,
                        "Uploaded": sn["publishedAt"][:10],
                        "ChCreated": ch.get("created", "")[:10] if ch.get("created") else "N/A",
                        "Country": country,
                        "Type": vtype,
                        "Duration": duration,
                        "DurationStr": f"{duration//60}:{duration%60:02d}",
                        "Faceless": "YES" if is_faceless else "MAYBE",
                        "FacelessScore": confidence,
                        "FacelessReasons": ", ".join(reasons) if reasons else "N/A",
                        "Keyword": kw,
                        "Thumb": sn["thumbnails"]["high"]["url"],
                        "Link": f"https://www.youtube.com/watch?v={vid}",
                        "ChannelLink": f"https://www.youtube.com/channel/{cid}"
                    }
                    
                    # Calculate growth potential
                    result_data["GrowthPotential"] = calculate_growth_potential(result_data)
                    
                    all_results.append(result_data)
    
    progress_bar.empty()
    status_text.empty()
    
    # STATS DISPLAY
    st.markdown("### 📊 Search Statistics")
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total Searched", stats["total_searched"])
    col2.metric("Passed Views", stats["passed_views"])
    col3.metric("Passed Subs", stats["passed_subs"])
    col4.metric("Passed Age", stats["passed_age"])
    col5.metric("Final Results", stats["final"])
    
    if not all_results:
        st.warning("😔 Kuch nahi mila! Try karo:")
        st.markdown("""
        - **Days** badha do (14 ya 30 days)
        - **Min Views** kam karo (5000 ya 1000)
        - **Channel Age** "Any" ya "2023" select karo
        - **Faceless Strictness** "Relaxed" karo
        - **Monetization Filter** disable karo
        - **More keywords** add karo
        """)
        st.stop()
    
    df = pd.DataFrame(all_results)
    df = df.sort_values("Views", ascending=False)
    df = df.drop_duplicates(subset="ChannelID", keep="first")
    df = df.reset_index(drop=True)
    
    st.success(f"🎉 **{len(df)} FACELESS VIRAL VIDEOS** mil gaye!")
    st.balloons()
    
    # SORTING
    st.markdown("### 🎯 Results")
    col1, col2 = st.columns(2)
    with col1:
        sort_by = st.selectbox("Sort By", ["Views", "Virality", "Engagement%", "Subs", "SubViewRatio", 
                                            "TotalVideos", "UploadsPerWeek", "MonetizationScore", 
                                            "GrowthPotential", "EstRevenue"])
    with col2:
        sort_order = st.selectbox("Order", ["Descending", "Ascending"])
    
    df = df.sort_values(by=sort_by, ascending=(sort_order == "Ascending"))
    
    # DISPLAY RESULTS
    for idx, r in df.iterrows():
        with st.container():
            st.markdown("---")
            col1, col2 = st.columns([3, 1])
            
            with col1:
                st.markdown(f"### {r['Title']}")
                
                st.markdown(
                    f"**📺 [{r['Channel']}]({r['ChannelLink']})** • "
                    f"👥 {r['Subs']:,} subs • "
                    f"🎬 **{r['TotalVideos']:,} videos** • "
                    f"👁️ {format_number(r['TotalChannelViews'])} total views • "
                    f"🌍 {r['Country']} • "
                    f"📂 **{r['Niche']}**"
                )
                
                st.markdown(
                    f"📅 **Created:** {r['ChCreated']} • "
                    f"⏰ **Upload Frequency:** {r['UploadSchedule']} • "
                    f"📊 {r['UploadsPerMonth']:.1f} videos/month"
                )
                
                # Monetization with Revenue
                if "LIKELY" in r['MonetizationStatus']:
                    st.success(f"💰 **{r['MonetizationStatus']}** (Score: {r['MonetizationScore']}%) | Est. Revenue: **${r['EstRevenue']:,.0f}** (${r['MonthlyRevenue']:,.0f}/mo)")
                elif "POSSIBLY" in r['MonetizationStatus']:
                    st.info(f"💰 **{r['MonetizationStatus']}** (Score: {r['MonetizationScore']}%) | Est. Revenue: **${r['EstRevenue']:,.0f}**")
                elif "CLOSE" in r['MonetizationStatus']:
                    st.warning(f"💰 **{r['MonetizationStatus']}** (Score: {r['MonetizationScore']}%)")
                else:
                    st.error(f"💰 **{r['MonetizationStatus']}** (Score: {r['MonetizationScore']}%)")
                
                with st.expander("📋 Monetization Details"):
                    reasons = r['MonetizationReasons'].split(" | ")
                    for reason in reasons:
                        st.markdown(f"- {reason}")
                
                # Stats
                col_a, col_b, col_c, col_d, col_e = st.columns(5)
                col_a.metric("👁️ Views", f"{r['Views']:,}")
                col_b.metric("🔥 Virality", f"{r['Virality']:,}/day")
                col_c.metric("💬 Engagement", f"{r['Engagement%']}%")
                col_d.metric("📈 Sub:View", f"{r['SubViewRatio']}x")
                col_e.metric("🚀 Growth", f"{r['GrowthPotential']}%")
                
                st.markdown(
                    f"⏱️ **Duration:** {r['DurationStr']} ({r['Type']}) • "
                    f"👍 {r['Likes']:,} likes • "
                    f"💬 {r['Comments']:,} comments • "
                    f"📤 Uploaded: {r['Uploaded']}"
                )
                
                if r['Faceless'] == "YES":
                    st.success(f"✅ **Faceless Channel** - Score: {r['FacelessScore']}% | {r['FacelessReasons']}")
                else:
                    st.info(f"🤔 Faceless Score: {r['FacelessScore']}% | {r['FacelessReasons']}")
                
                st.markdown(f"🔑 Keyword: `{r['Keyword']}`")
                st.markdown(f"[▶️ Watch Video]({r['Link']})")
            
            with col2:
                st.image(r["Thumb"], use_container_width=True)
                st.markdown("---")
                st.markdown(f"**📊 Channel Stats**")
                st.markdown(f"🎬 {r['TotalVideos']} videos")
                st.markdown(f"📅 {r['UploadsPerWeek']:.1f}/week")
                st.markdown(f"💰 ${r['MonthlyRevenue']:,.0f}/mo")
                st.markdown(f"🚀 {r['GrowthPotential']}% growth")
    
    # ------------------------------------------------------------
    # EXPORT SECTION
    # ------------------------------------------------------------
    st.markdown("---")
    st.markdown("### 📥 Download Reports")
    
    export_cols = st.columns(4)
    
    # CSV Export
    with export_cols[0]:
        csv = df.to_csv(index=False).encode('utf-8')
        st.download_button(
            "📥 CSV",
            data=csv,
            file_name=f"faceless_viral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    # Excel Export
    with export_cols[1]:
        if "Excel" in export_format:
            excel_data = generate_excel_report(df)
            if excel_data:
                st.download_button(
                    "📥 Excel",
                    data=excel_data,
                    file_name=f"faceless_viral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.info("Install openpyxl for Excel")
    
    # HTML Report
    with export_cols[2]:
        if "HTML Report" in export_format:
            html_report = generate_html_report(df, stats)
            st.download_button(
                "📥 HTML Report",
                data=html_report,
                file_name=f"faceless_viral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html",
                mime="text/html",
                use_container_width=True
            )
    
    # PDF Report
    with export_cols[3]:
        if "PDF Report" in export_format:
            pdf_data = generate_pdf_report(df, stats)
            if pdf_data:
                st.download_button(
                    "📥 PDF Report",
                    data=pdf_data,
                    file_name=f"faceless_viral_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
            else:
                st.info("Install fpdf for PDF: pip install fpdf")
    
    # Summary Stats
    st.markdown("### 📈 Summary Statistics")
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.markdown("**💰 Monetization**")
        monetized_count = len(df[df['MonetizationScore'] >= 70])
        possibly_count = len(df[(df['MonetizationScore'] >= 50) & (df['MonetizationScore'] < 70)])
        not_monetized = len(df[df['MonetizationScore'] < 50])
        st.markdown(f"- 🟢 Monetized: {monetized_count}")
        st.markdown(f"- 🟡 Possibly: {possibly_count}")
        st.markdown(f"- 🔴 Not Yet: {not_monetized}")
    
    with col2:
        st.markdown("**📅 Upload Frequency**")
        daily = len(df[df['UploadsPerWeek'] >= 7])
        active = len(df[(df['UploadsPerWeek'] >= 3) & (df['UploadsPerWeek'] < 7)])
        regular = len(df[(df['UploadsPerWeek'] >= 1) & (df['UploadsPerWeek'] < 3)])
        st.markdown(f"- 🔥 Daily+: {daily}")
        st.markdown(f"- 📈 Active: {active}")
        st.markdown(f"- ✅ Regular: {regular}")
    
    with col3:
        st.markdown("**📂 Top Niches**")
        niche_counts = df['Niche'].value_counts().head(4)
        for niche, count in niche_counts.items():
            st.markdown(f"- {niche}: {count}")
    
    with col4:
        st.markdown("**💵 Revenue Est.**")
        st.markdown(f"- Total: ${df['EstRevenue'].sum():,.0f}")
        st.markdown(f"- Avg: ${df['EstRevenue'].mean():,.0f}")
        st.markdown(f"- Top: ${df['EstRevenue'].max():,.0f}")
    
    with col5:
        st.markdown("**🌍 Countries**")
        country_counts = df['Country'].value_counts().head(4)
        for country, count in country_counts.items():
            st.markdown(f"- {country}: {count}")
    
    # Table View
    with st.expander("📋 View All Data (Table)"):
        st.dataframe(
            df[["Title", "Channel", "Views", "Virality", "Subs", "TotalVideos", 
                "UploadsPerWeek", "MonetizationStatus", "MonetizationScore",
                "EstRevenue", "GrowthPotential", "Niche", "Country", "Faceless"]],
            use_container_width=True,
            height=400
        )

# FOOTER
st.markdown("---")
st.caption("Made with ❤️ for Muhammed Rizwan Qamar | Faceless Viral Hunter PRO 2025")
